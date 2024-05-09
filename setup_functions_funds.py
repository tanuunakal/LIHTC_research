from tqdm import tqdm
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
import urllib
import warnings
import json
import re
from google.colab import drive

class FundsSetupFunctions:
    def get_sheet_name_funds(self, year):
        if int(year) > 2020:
            return "Part II-Sources of Funds"
        else:
            return "Part III-Sources of Funds"


    def get_user_selection(self, section_name):
        if section_name == "CONSTRUCTION FINANCING":
            return {"Mortgage A", "Mortgage B", "Mortgage C",
                    "Federal Grant", "State, Local, or Private Grant", "Deferred Developer Fees", "Federal Housing Credit Equity",
                    "State Housing Credit Equity", "Other Type (specify)", "Total Construction Period Costs from Development Budget:"}
        if section_name == "PERMANENT FINANCING":
            return {"Mortgage A (Lien Position 1)", "Mortgage B (Lien Position 2)", "Mortgage C (Lien Position 3)",
                    "Other:", "Foundation, charity or other govt*", "Deferred Devlpr Fee", "Federal Grant",
                    "State, Local, or Private Grant", "Federal Housing Credit Equity", "State Housing Credit Equity",
                    "Historic Credit Equity", "Invstmt Earnings: T-E Bonds", "Invstmt Earnings: Taxable Bonds",
                    "Income from Operations", "Total Permanent Financing:"}
        return {}

    def get_boundaries(self, f_id, section_name):
        sections_file = '/content/drive/MyDrive/lihtc/sources_of_funds_sections.json'
        with open(sections_file, 'r') as file:
            sections = json.load(file)
        return sections[f_id][section_name]


    #get all the entitites within a row
    # iterate through all the cells in the row and get all the colored ones after the value
    def get_entities_for_selections(self, ws,row_num, row, user_selections):
        data = []
        selection_row = {}
        key = None
        for cell in ws[row_num]:
            if cell.value in user_selections:
                key = cell.value
                user_selections.remove(key)
            if cell.fill.start_color.index == 41 and key:
                data.append(cell.value)
            if cell.fill.start_color.index == 'FFCCFFFF' and key:
                data.append(row[cell.column - 1])
            if cell.fill.start_color.index == 'FFCCFFCC' and key:
                data.append(row[cell.column - 1])
            if cell.fill.start_color.index == 42 and key:
                data.append(row[cell.column - 1])
            if cell.font.bold and key and key == 'Total Permanent Financing:':
                data.append(row[cell.column - 1])
        selection_row[key] = data
        return selection_row

    def get_data_for_sheet (self, wb, sheet_name, user_selections,
                            section_start_row, section_end_row):
        ws = wb[sheet_name]
        data = {}
        for row_num, row in enumerate(ws.iter_rows(min_row=section_start_row, \
                                                    max_row=section_end_row, \
                                                    values_only=True),
                                        start=section_start_row):
            if user_selections.intersection(row):
                data.update(self.get_entities_for_selections(ws, row_num, row, user_selections))
        return data


    def get_data_by_drive_id(self, drive_id, sheet_name, year,
                            boundaries, section_name):
        wb = load_workbook_from_url (construct_drive_url (drive_id))
        user_selections = self.get_user_selection(section_name)
        data = self.get_data_for_sheet (wb, sheet_name, user_selections, boundaries[0],
                                    boundaries[1])
        if section_name == "CONSTRUCTION FINANCING":
            self.standardize_dict_comstruction_funds(data)
        if section_name == "PERMANENT FINANCING":
            self.standardize_dict_permanent_funds(data)
        return data



    # @title ### Standardization functions
    def standardize_dict_comstruction_funds(self, data):
        for key, val in data.items():
            if len(val) == 4:
                dict_vals = {}
                dict_vals["Name of Financing Entity"] = val[0]
                dict_vals["Amount"] = val[1]
                dict_vals["Effective Interest Rate"] = val[2]
                dict_vals["Term (In Months)"] = val[3]
                data[key] = dict_vals
            if len(val) == 2:
                dict_vals = {}
                dict_vals["Name of Financing Entity"] = val[0]
                dict_vals["Amount"] = val[1]
                data[key] = dict_vals
            if len(val) == 3:
                dict_vals = {}
                dict_vals["Type"] = val[0]
                dict_vals["Name of Financing Entity"] = val[1]
                dict_vals["Amount"] = val[2]
                data[key] = dict_vals
            if len(val) == 1:
                data[key] = val[0]

    def standardize_dict_permanent_funds(self, data):
        for key, val in data.items():
            if len(val) == 7:
                dict_vals = {}
                dict_vals["Name of Financing Entity"] = val[0]
                dict_vals["Amount"] = val[1]
                dict_vals["Effective Interest Rate"] = val[2]
                dict_vals["Term (Years)"] = val[3]
                dict_vals["Amort. (Years)"] = val[4]
                dict_vals["Loan Type"] = val[5]
                dict_vals["Annual Debt Service in Year One"] = val[6]
                data[key] = dict_vals
            if len(val) == 8:
                dict_vals = {}
                dict_vals["Type"] = val[0]
                dict_vals["Name of Financing Entity"] = val[1]
                dict_vals["Amount"] = val[2]
                dict_vals["Effective Interest Rate"] = val[3]
                dict_vals["Term (Years)"] = val[4]
                dict_vals["Amort. (Years)"] = val[5]
                dict_vals["Loan Type"] = val[6]
                dict_vals["Annual Debt Service in Year One"] = val[7]
                data[key] = dict_vals
            if len(val) == 2:
                dict_vals = {}
                dict_vals["Name of Financing Entity"] = val[0]
                dict_vals["Amount"] = val[1]
                data[key] = dict_vals
            if len(val) == 3:
                dict_vals = {}
                dict_vals["Type"] = val[0]
                dict_vals["Name of Financing Entity"] = val[1]
                dict_vals["Amount"] = val[2]
                data[key] = dict_vals
            if len(val) == 1:
                data[key] = val[0]

