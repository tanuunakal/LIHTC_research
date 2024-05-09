from tqdm import tqdm
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
import urllib
import warnings
import json
import re
from setup_functions import SetupFunctions

class SubheadingFunctions:
    
    
    def __init__(self):
        self.setup_instance = SetupFunctions()

    def list_to_dict_with_count(self, lst):
        counts = {}
        for item in lst:
            counts[item] = counts.get(item, 0) + 1
        return counts

    def get_sheet_name_funds(self, year):
        if int(year) > 2020:
            return "Part II-Sources of Funds"
        else:
            return "Part III-Sources of Funds"

    def get_all_files_bolded(self, drive_dir_id):
        data = []
        for file_info in tqdm(self.setup_instance.folder_list (drive_dir_id)):
            year_str = self.setup_instance.get_year_by_name(file_info[1])
            print(year_str)
            yearly_data = self.get_data_for_year(file_info[0], self.get_sheet_name_funds(year_str))
            data = data + yearly_data
        return data


    #get data for the year for all the fields in the section
    def get_data_for_year(self, drive_dir_id, sheet_name):
        data = []
        for file_info in tqdm(self.setup_instance.folder_list (drive_dir_id)):
            bolded_data = self.get_bolded_by_drive_id(file_info[0], sheet_name)
            data = data + bolded_data
        return data

    def get_bolded(self, wb, sheet_name):
        ws = wb[sheet_name]
        headings = []
        for row_num, row in enumerate(ws.iter_rows(values_only=True, max_col=5),
                                        start=1):
            # Get the entity from the cell next to user_selection that has a blue fill
            for cell in ws[row_num]:
                if cell.font.bold and cell.value:
                    headings.append(cell.value)
        return headings

    def get_bolded_by_drive_id(self, drive_id, sheet_name):
        wb = self.setup_instance.load_workbook_from_url (self.setup_instance.construct_drive_url(drive_id))
        data = self.get_bolded (wb, sheet_name)
        print(data)
        return data

