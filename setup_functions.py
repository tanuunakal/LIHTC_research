import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
import urllib
import warnings
import json
import re


def construct_drive_url (file_id):
  return f"https://drive.google.com/uc?export=download&id={file_id}"

def load_workbook_from_url(url):
  """ https://stackoverflow.com/a/64725882 """
  file = urllib.request.urlopen(url).read()
  return load_workbook(filename = BytesIO(file), data_only=True)

#The name of the sheet changes according to the year,
# this function retrieves the sheet_name by year
def get_sheet_name(year):
  if int(year) > 2020:
    return "Part I-Project Identification"
  else:
    return "Part I-Project Information"

def get_user_selection(section_name):
  if section_name == "Applicant Contact":
    return {"Organization Name", "Address", "City",
            "State", "Zip+4", "Contact", "Office Phone",
            "E-mail", "Name"}
  if section_name == "Project Location":
    return {"Project Name", "City", "County", "Acreage",
            "Site Acreage", "In USDA Rural Area?",
            "Political Jurisdiction", "Zip+4", "Name of Chief Elected Official",
            "Title", "Site Geo Coordinates", "Longitude:", "Site Geo Coordinates     (##.######)"}
  return {}


#list all [ids, name] of files directly under folder folder_id
def folder_list(folder_id):
  from googleapiclient.discovery import build
  gdrive = build('drive', 'v3').files()
  res = gdrive.list(q="'%s' in parents" % folder_id).execute()
  return [[f['id'], f['name']] for f in res['files']]

def get_boundaries(year, section_name):
  sections_file = '/content/drive/MyDrive/lihtc/sections.json'
  with open(sections_file, 'r') as file:
    sections = json.load(file)
  return sections[year][section_name]

#gets the year given the folder name, eg:
def get_year_by_name(folder_name):
  return re.findall(r'\d{4}$', folder_name)[0]

def get_entities_for_selections(ws,row_num, row, user_selections):
  data = {}
  key = None
  for cell in ws[row_num]:
    if cell.value in user_selections:
      key = cell.value
      user_selections.remove(key)
    if cell.fill.start_color.index == 41 and key:
      data[key] = cell.value
      key = None
    if cell.fill.start_color.index == 42 and key:
      data[key] = row[cell.column - 1]
      key = None
  return data

def get_data_for_sheet (wb, sheet_name, user_selections,
                        section_start_row, section_end_row):
  ws = wb[sheet_name]
  data = {}
  is_search_section = False
  for row_num, row in enumerate(ws.iter_rows(min_row=section_start_row, \
                                             max_row=section_end_row, \
                                             values_only=True),
                                start=section_start_row):
    if user_selections.intersection(row):
      data.update(get_entities_for_selections(ws, row_num, row, user_selections))
  return data

def standardize_dict_project_location(user_selections, data):
  if "Name of Chief Elected Official" in user_selections:
    data["Name of Chief Elected Official"] = ""
  if "Political Jurisdiction" in user_selections:
    data["Political Jurisdiction"] = ""
  if "Zip+4" in user_selections:
    data["Zip+4"] = ""
  if "Title" in user_selections:
    data["Title"] = ""
  if "Longitude:" not in user_selections:
    if "Site Geo Coordinates" in data:
      data["Site Geo Coordinates"] = str(data["Site Geo Coordinates"]) \
                                  + ", " + str(data["Longitude:"])
    elif  "Site Geo Coordinates     (##.######)" in data:
      data["Site Geo Coordinates"] = str(data["Site Geo Coordinates     (##.######)"]) \
                                  + ", " + str(data["Longitude:"])
      del data["Site Geo Coordinates     (##.######)"]
    del data["Longitude:"]
  if "Acreage" in user_selections:
    data["Acreage"] = data["Site Acreage"]
    del data["Site Acreage"]

def standardize_dict_applicant_contact(user_selections, data):
  if "Organization Name" in user_selections:
    data["Organization Name"] = ""
  if "Contact" in user_selections:
    data["Contact"] = data["Name"]
    del data["Name"]

def get_data_by_drive_id(drive_id, sheet_name, year,
                         boundaries, section_name):
  wb = load_workbook_from_url (construct_drive_url (drive_id))
  user_selections = get_user_selection(section_name)
  data = get_data_for_sheet (wb, sheet_name, user_selections, boundaries[0],
                             boundaries[1])
  if section_name == "Applicant Contact":
    standardize_dict_applicant_contact(user_selections, data)
  if section_name == "Project Location":
    standardize_dict_project_location(user_selections, data)
  return data

#get data for the year for all the fields in the section
def get_data_for_year(drive_dir_id, sheet_name,
                      year, section_name):
  data = []
  for file_info in tqdm(folder_list (drive_dir_id)):
    f_id = file_info[0]
    f_name = file_info[1]
    print("f name is: " + f_name)
    boundaries = get_boundaries(year, section_name)
    if f_id in weird_ids and section_name == "Project Location":
      boundaries = [25, 44]
    data_file = get_data_by_drive_id(file_info[0], sheet_name,
                                     year, boundaries, section_name)
    data_file["gdrive_id"] = f_id
    data_file["file_id"] = f_name
    data.append(data_file)
  return data

# write all files for all years
def write_all_files_to_json(drive_dir_id, section_name):
  data = []
  for file_info in tqdm (folder_list (drive_dir_id)):
    year_str = get_year_by_name(file_info[1])
    print(year_str)
    yearly_data = get_data_for_year(file_info[0], get_sheet_name(year_str),
                                    year_str, section_name)
    data = data + yearly_data
  with open(json_file_path, 'w') as json_file:
   json.dump(data, json_file)

