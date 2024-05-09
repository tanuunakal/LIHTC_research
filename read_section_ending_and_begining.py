from tqdm import tqdm
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
import urllib
import warnings
import json
import re
from setup_functions import SetupFunctions

class SectionFunctions:
    
    
    def __init__(self):
        self.setup_instance = SetupFunctions()
    
    def write_sections_by_year(self, drive_dir_id):
        sections = {}
        for file_info in tqdm(self.setup_instance.folder_list (drive_dir_id)):
            year = int(self.setup_instance.get_year_by_name(file_info[1]))
            yearly_section = self.get_section_for_year(file_info[0], self.setup_instance.get_sheet_name(year))
            sections[year] = yearly_section
        with open(sections_file, 'w') as file:
            json.dump(sections, file)

    def get_section_for_year(self, drive_id, sheet_name):
        my_instance = MyFunctions()
        file_of_year_info = self.setup_instance.folder_list (drive_id)[0]
        file_id = file_of_year_info[0]
        wb = self.setup_instance.load_workbook_from_url (self.setup_instance.construct_drive_url (file_id))
        ws = wb[sheet_name]
        return self.get_sections_for_sheet(ws)

    def get_sections_for_sheet(self, ws):
        sections = {}
        is_search_section = False

        for row_num, row in enumerate(ws.iter_rows(values_only=True),
                                        start=1):
            if self.is_start_of_contact_section(row):
                sections["Applicant Contact"] = [row_num]
            if self.is_end_of_contact_section(row):
                sections["Applicant Contact"].append(row_num)
            if self.is_start_of_project_location_section(row):
                sections["Project Location"] = [row_num]
            if self.is_end_of_project_location_section(row):
                sections["Project Location"].append(row_num)

        return sections

    def is_start_of_contact_section(self, row):
        return ("APPLICANT CONTACT FOR APPLICATION SUBMISSION AND REVIEW" in row) \
        or ("APPLICANT CONTACT FOR APPLICATION REVIEW" in row)

    def is_end_of_contact_section(self, row):
        return "PROJECT LOCATION" in row

    def is_start_of_project_location_section(self, row):
        return "PROJECT LOCATION" in row

    def is_end_of_project_location_section(self, row):
        return ("PROJECT DESCRIPTION" in row) \
        or ("WAIVERS AND/OR PRE-APPROVALS " in row)

